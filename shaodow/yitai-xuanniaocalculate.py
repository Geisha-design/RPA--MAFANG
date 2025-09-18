import configparser
import calendar
import datetime
import json
import smtplib
import time
from builtins import list
from email.header import Header
from email.mime.application import MIMEApplication
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
import requests
# import xlwt
from datetime import timedelta
from openpyxl import load_workbook
# import requests
import pandas as pd
from openpyxl import Workbook
import pymysql
# import xlwt
from sqlalchemy.sql.functions import now

# 用于进行Excel转换的信息结成队列

# 小笨鸟拖车
listSeqA = []
# 小笨鸟报关
listSeqB = []
# 大笨鸟报关
listSeqC = []
# 大笨鸟拖车
listSeqD = []
# 大笨鸟操作票数
listSeqE = []

now = datetime.date.today()
this_year_start = str(datetime.datetime(now.year, 1, 1))
this_month_start = str(datetime.datetime(now.year, now.month, 1))
this_month_end = str(datetime.datetime(now.year, now.month, calendar.monthrange(now.year, now.month)[1]))

this_last_month_start = str(datetime.datetime(now.year, now.month - 1, 1))
this_month_end = str(datetime.datetime(now.year, now.month, calendar.monthrange(now.year, now.month)[1]))


def idCapture(companies):
    user = 'uct'
    passwd = 'Uct_0828'
    port = 3306
    db = 'uct'
    host = "rm-uf6pu8262p430c5z5.mysql.rds.aliyuncs.com"
    conn = pymysql.connect(host=host
                           , user=user
                           , passwd=passwd
                           , port=port
                           , db=db
                           , charset='utf8'
                           )
    cur = conn.cursor()
    try:
        listIds = []
        for i in range(len(companies)):
            gongsicanshu = companies[i]
            sql = """
            SELECT id FROM `uct_org` where `name` in (%(gongsi)s)
            """
            cur.execute(sql, {"gongsi": gongsicanshu})
            myresult = cur.fetchall()
            for x in myresult:
                listIds.append(x[0])
        return listIds
    except Exception as e:
        print(e)
        conn.rollback()
        conn.close()


def xiaobenniaotuoche(theids, timeFlagA, timeFlagB):
    user = 'ifsclient'
    passwd = 'Ifsclient_0330'
    port = 3306
    db = 'ifsclient'
    host = "rm-uf6l1kfllh0d8w0k9.mysql.rds.aliyuncs.com"
    # 连接数据库
    conn = pymysql.connect(host=host  # 连接名称，默认127.0.0.1
                           , user=user  # 用户名
                           , passwd=passwd  # 密码
                           , port=port  # 端口，默认为3306
                           , db=db  # 数据库名称
                           , charset='utf8'  # 字符编码
                           )
    cur = conn.cursor()  # 生成游标对象
    try:
        # sql = "insert into ocean_flight_etd_eta3  values(%s,%s,%s,%s,%s)"
        # # sql = "insert into" + table1 + "values(%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)"
        # param = (a,b,c,d,e)

        # mycursor = mydb.cursor()
        # sql = """
        # SELECT ORG_ID,count(1) FROM lit_booking_declare_order WHERE DELETED_FLAG = '0' and declare_way = '0' and org_id in (%(theid)s) and create_time >= '2024-03-01 00:00:00' and create_time < '2023-09-01 00:00:00' GROUP BY ORG_ID;
        # """

        sql = """
        SELECT org_id, count(1) FROM  ifs_trailer_trust_order WHERE state IN (0, 1)  AND deleted_flag = 0 and org_id in (%(theid)s) and create_time >= (%(timeFlagA)s) and create_time < (%(timeFlagB)s) GROUP BY org_id;
        """
        # cur.execute("SELECT * FROM ocean_flight_etd_eta3")
        cur.execute(sql, {"theid": theids, "timeFlagA": timeFlagA, "timeFlagB": timeFlagB})
        # cur.execute("")
        myresult = cur.fetchall()

        # for x in myresult:
        #     print(x)
        #     if(x==None or x==''):
        #         print(0)
        if myresult:
            for x in myresult:
                print(x[1])  # 打印每条记录
                listSeqA.append(x[1])
        else:
            # print("No data found.")
            print(0)
            listSeqA.append(0)
        # cur.execute(sql, param)
        # conn.commit() # 提交到数据库执行
    except Exception as e:
        print(e)
        conn.rollback()  # 如果发生错误则回滚
        conn.close()  # 关闭数据库连接


def xiaobenniaobaoguan(theids, timeFlagA, timeFlagB):
    user = 'ifsclient'
    passwd = 'Ifsclient_0330'
    port = 3306
    db = 'ifsclient'
    host = "rm-uf6l1kfllh0d8w0k9.mysql.rds.aliyuncs.com"
    # 连接数据库
    conn = pymysql.connect(host=host  # 连接名称，默认127.0.0.1
                           , user=user  # 用户名
                           , passwd=passwd  # 密码
                           , port=port  # 端口，默认为3306
                           , db=db  # 数据库名称
                           , charset='utf8'  # 字符编码
                           )
    cur = conn.cursor()  # 生成游标对象
    try:
        # sql = "insert into ocean_flight_etd_eta3  values(%s,%s,%s,%s,%s)"
        # # sql = "insert into" + table1 + "values(%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)"
        # param = (a,b,c,d,e)
        # mycursor = mydb.cursor()
        # sql = """
        # SELECT ORG_ID,count(1) FROM lit_booking_declare_order WHERE DELETED_FLAG = '0' and declare_way = '0' and org_id in (%(theid)s) and create_time >= '2024-03-01 00:00:00' and create_time < '2023-09-01 00:00:00' GROUP BY ORG_ID;
        # """
        sql = """
        SELECT org_id, count(1) FROM `ifs_declaration_order` WHERE order_status_code not in('5', '2', '7') AND deleted_flag = 0 and org_id in (%(theid)s) and create_time >= (%(timeFlagA)s) and create_time < (%(timeFlagB)s) AND source_flag =2 GROUP BY org_id;
        """
        # cur.execute("SELECT * FROM ocean_flight_etd_eta3")
        cur.execute(sql, {"theid": theids, "timeFlagA": timeFlagA, "timeFlagB": timeFlagB})
        # cur.execute("")
        myresult = cur.fetchall()

        if myresult:
            for x in myresult:
                print(x[1])  # 打印每条记录
                listSeqB.append(x[1])
        else:
            # print("No data found.")
            print(0)
            listSeqB.append(0)
        # cur.execute(sql, param)
        # conn.commit() # 提交到数据库执行
    except Exception as e:
        print(e)
        conn.rollback()  # 如果发生错误则回滚
        conn.close()  # 关闭数据库连接


# 大笨鸟报关
def dabenniaobaoguan(theids, timeFlagA, timeFlagB):
    user = 'litsaas'
    passwd = 'Litsaas_0607'
    port = 3306
    db = 'litsaas'
    host = "rm-uf6160k88d3snm052.mysql.rds.aliyuncs.com"
    # 连接数据库
    conn = pymysql.connect(host=host  # 连接名称，默认127.0.0.1
                           , user=user  # 用户名
                           , passwd=passwd  # 密码
                           , port=port  # 端口，默认为3306
                           , db=db  # 数据库名称
                           , charset='utf8'  # 字符编码
                           )
    cur = conn.cursor()  # 生成游标对象
    try:
        # mycursor = mydb.cursor()
        sql = """
                SELECT ORG_ID,count(1) FROM lit_booking_declare_order WHERE DELETED_FLAG = '0' and declare_way = '0' and org_id in (%(theid)s) and create_time >=  (%(timeFlagA)s) and create_time < (%(timeFlagB)s) GROUP BY ORG_ID;
        """
        # SELECT org_id, count(1) FROM `ifs_declaration_order` WHERE order_status_code not in('5', '2', '7') AND deleted_flag = 0 and org_id in (%(theid)s) and create_time >= '2024-09-01 00:00:00' and create_time < '2024-10-01 00:00:00' AND source_flag in(41,42) GROUP BY org_id;

        # cur.execute("SELECT * FROM ocean_flight_etd_eta3")
        cur.execute(sql, {"theid": theids, "timeFlagA": timeFlagA, "timeFlagB": timeFlagB})
        # cur.execute("")
        myresult = cur.fetchall()

        # 检查是否有结果
        if myresult:
            for x in myresult:
                print(x[1])  # 打印每条记录
                listSeqC.append(x[1])
        else:
            # print("No data found.")
            print(0)
            listSeqC.append(0)

        # for x in myresult:

        #     print(x)
        #     if(x==None or x=='' or x==[] or x==N):
        #         print(0)

        # cur.execute(sql, param)
        # conn.commit() # 提交到数据库执行
    except Exception as e:
        print(e)
        conn.rollback()  # 如果发生错误则回滚
        conn.close()  # 关闭数据库连接


# 大笨鸟拖车
def dabenniaotuoche(saika, timeFlagA, timeFlagB):
    user = 'litsaas'
    passwd = 'Litsaas_0607'
    port = 3306
    db = 'litsaas'
    host = "rm-uf6160k88d3snm052.mysql.rds.aliyuncs.com"
    # 连接数据库
    conn = pymysql.connect(host=host  # 连接名称，默认127.0.0.1
                           , user=user  # 用户名
                           , passwd=passwd  # 密码
                           , port=port  # 端口，默认为3306
                           , db=db  # 数据库名称
                           , charset='utf8'  # 字符编码
                           )
    cur = conn.cursor()  # 生成游标对象
    try:
        # sql = "insert into ocean_flight_etd_eta3  values(%s,%s,%s,%s,%s)"
        # # sql = "insert into" + table1 + "values(%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)"
        # param = (a,b,c,d,e)
        # mycursor = mydb.cursor()
        # cur.execute("SELECT * FROM ocean_flight_etd_eta3")
        # cur.execute("SELECT ORG_ID,count(1) FROM lit_trailer_trust_order WHERE DELETED_FLAG = '0' and trailer_type = '0' and org_id in ('100395921','100396182') and create_time >= '2023-06-01 00:00:00' and create_time < '2023-07-01 00:00:00' GROUP BY ORG_ID;")

        sql = "SELECT ORG_ID,count(1) FROM lit_trailer_trust_order WHERE DELETED_FLAG = '0' and trailer_type = '0' and org_id in (%(tuoche)s) and create_time >= (%(timeFlagA)s) and create_time < (%(timeFlagB)s) GROUP BY ORG_ID;"
        # cur.execute("SELECT ORG_ID,count(1) FROM lit_trailer_trust_order WHERE DELETED_FLAG = '0' and trailer_type = '0' and org_id in (%(tuoche)s) and create_time >= '2023-06-01 00:00:00' and create_time < '2023-07-01 00:00:00' GROUP BY ORG_ID;")
        # sql = """
        # SELECT ORG_ID,count(1) FROM lit_booking_declare_order WHERE DELETED_FLAG = '0' and declare_way = '0' and org_id in (%(theid)s) and create_time >= '2023-08-01 00:00:00' and create_time < '2023-09-01 00:00:00' GROUP BY ORG_ID;
        # """
        # cur.execute("SELECT * FROM ocean_flight_etd_eta3")
        cur.execute(sql, {"tuoche": saika, "timeFlagA": timeFlagA, "timeFlagB": timeFlagB})

        myresult = cur.fetchall()
        # for x in myresult:
        #     print(x)
        #     if(x==None or x==''):
        #         print(0)
        if myresult:
            for x in myresult:
                print(x[1])  # 打印每条记录
                listSeqD.append(x[1])
        else:
            # print("No data found.")
            print(0)
            listSeqD.append(0)

        # cur.execute(sql, param)
        # conn.commit() # 提交到数据库执行
    except Exception as e:
        print(e)
        conn.rollback()  # 如果发生错误则回滚
        conn.close()  # 关闭数据库连接


# 大笨鸟操作票数
def dabenniaocaozuopiaoshu(saika, timeFlagA, timeFlagB):
    user = 'litsaas'
    passwd = 'Litsaas_0607'
    port = 3306
    db = 'litsaas'
    host = "rm-uf6160k88d3snm052.mysql.rds.aliyuncs.com"
    # 连接数据库
    conn = pymysql.connect(host=host  # 连接名称，默认127.0.0.1
                           , user=user  # 用户名
                           , passwd=passwd  # 密码
                           , port=port  # 端口，默认为3306
                           , db=db  # 数据库名称
                           , charset='utf8'  # 字符编码
                           )
    cur = conn.cursor()  # 生成游标对象
    try:
        # sql = "insert into ocean_flight_etd_eta3  values(%s,%s,%s,%s,%s)"
        # # sql = "insert into" + table1 + "values(%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)"
        # param = (a,b,c,d,e)

        # mycursor = mydb.cursor()

        # sql = "SELECT ORG_ID,count(1) FROM lit_trailer_trust_order WHERE DELETED_FLAG = '0' and trailer_type = '0' and org_id in (%(tuoche)s) and create_time >= '2024-03-01 00:00:00' and create_time < '2023-09-01 00:00:00' GROUP BY ORG_ID;"
        # cur.execute("SELECT ORG_ID,count(1) FROM lit_trailer_trust_order WHERE DELETED_FLAG = '0' and trailer_type = '0' and org_id in (%(tuoche)s) and create_time >= '2023-06-01 00:00:00' and create_time < '2023-07-01 00:00:00' GROUP BY ORG_ID;")
        sql = "select org_id, sum(num) from ( (SELECT ORG_ID,count(1) as num FROM lit_booking_declare_order WHERE DELETED_FLAG = '0' and org_id in (%(caozuopiaoshu)s)and create_time >= (%(timeFlagA)s) and create_time < (%(timeFlagB)s) GROUP BY ORG_ID)union ALL(SELECT ORG_ID,count(1) as num FROM lit_trailer_trust_order WHERE DELETED_FLAG = '0' and org_id in (%(caozuopiaoshu)s) and create_time >= (%(timeFlagA)s) and create_time < (%(timeFlagB)s) GROUP BY ORG_ID)union ALL(SELECT ORG_ID,count(1) as num FROM lit_job_head WHERE DELETED_FLAG = '0'  and status='0' and org_id in (%(caozuopiaoshu)s) and create_time >= (%(timeFlagA)s) and create_time < (%(timeFlagB)s) GROUP BY ORG_ID)) as a group by org_id"
        cur.execute(sql, {"caozuopiaoshu": saika, "timeFlagA": timeFlagA, "timeFlagB": timeFlagB})
        myresult = cur.fetchall()
        # for x in myresult:
        #     print(x)
        #     if(x==None or x==''):
        #         print(0)
        if myresult:
            for x in myresult:
                print(x[1])  # 打印每条记录  列表第二个才是真正需要获取的数据
                listSeqE.append(x[1])
        else:
            # print("No data found.")
            print(0)
            listSeqE.append(0)

    except Exception as e:
        print(e)
        conn.rollback()  # 如果发生错误则回滚
        conn.close()  # 关闭数据库连接


companies = [
    "宁波民亨国际物流有限公司",
    "宁波斯必达供应链管理有限公司",
    "宁波鼎领国际物流有限公司",
    "宁波鼎领国际物流有限公司上海分公司",
    "浙江湖州华凯国际货运代理有限公司",
    "浙江华凯供应链管理有限公司",
    "湖州陆港国际供应链有限公司",
    "浙江湖州港通供应链科技有限公司",
    "浙江德清港泰物流有限公司",
    "浙江长川物流有限公司",
    "浙江泽惠供应链管理有限公司",
    "浙江洋帆供应链管理有限公司",
    "浙江多多马国际货运代理有限公司",
    "浙江凤起物流有限公司",
    "湖州振和供应链管理有限公司",
    "诸暨市中创物流有限公司",
    "深圳拓普斯达国际物流有限公司",
    "上海拓普斯达国际物流有限公司",
    "成都拓普斯达国际物流有限公司",
    "台湾拓達國際有限公司",
    "俄罗斯拓普斯达国际物流有限公司",
    "上海丝达进出口有限公司",
    "香港丝达有限公司",
    "深圳奇点国际贸易有限公司",
    "浙江锦鸿物流有限公司",
    "沃泰国际物流宁波有限公司",
    "湖州易航国际货运代理有限公司",
    "宁波亦豪供应链管理有限公司",
    "杭州环宇国际货运有限公司",
    "上海市食品进出口国际货运有限公司",
    "浙江福瑞鑫供应链服务有限公司",
    "浙江运领船务有限公司",
    "宁波乐洲国际物流有限公司",
    "领航广州国际物流有限公司",
    "义乌市金隆义境通供应链服务有限公司",
    "义乌市桐利国际货运代理有限公司",
    "义乌市泊胜国际货运代理有限公司",
    "大连锦程国际物流有限公司宁波分公司",
    "宁波众拾吉运国际货运代理有限公司",
    "义乌市天诚国际货运代理有限公司",
    "杭州高新供应链管理有限公司",
    "深圳市中传国际物流有限公司",
    "宁波维亚国际货运代理有限公司",
    "锐链宁波科技有限公司",
    "江苏锦晟国际货运代理有限公司",
    "江苏集信国际货运代理有限公司",
    "浙江正彧供应链管理有限公司",
    "杭州至星国际货运代理有限公司",
    "浙江中星供应链管理有限公司",
    "浙江华迅国际物流有限公司",
    "江苏清源国际货运代理有限公司",
    "浙江华威国际集装箱运输有限公司宁波分公司",
    "宁波鸿远国际物流有限公司",
    "浙江佰度物流有限公司"
]

company_list = [
    "大连锦程国际贸易有限公司",
    "宁波合羽国际贸易有限公司",
    "宁波致翔供应链科技有限公司",
    "青岛捷瀛船务有限公司",
    "宁波全顺捷运物流有限公司",
    "宁波速恒国际物流有限公司",
    "宁波至华国际物流有限公司",
    "宁波中景新程物流集团有限公司",
    "中集世联达亚联国际物流有限公司宁波分公司",
    "浙江明骏供应链管理有限公司",
    "浙江天时利蜻蜓供应链科技有限公司",
    "浙江泽惠供应链管理有限公司",
    "五朵金花运营中心",
    "大连锦程国际物流有限公司",
    "浙江某某国际货运代理有限公司 (演示)",
    "上海龙辕供应链管理有限公司",
    "上海旭华物流有限公司宁波分公司",
    "宁波晴海国际物流有限公司",
    "宁波鼎领国际物流有限公司",
    "深圳凯高现代物流有限公司",
    "安必快外贸综合服务（惠州）有限公司",
    "浙江佳福国际货运代理有限公司",
    "江苏常利来国际货运代理有限公司",
    "联邦快递（中国）有限公司深圳分公司",
    "江苏明美国际物流有限公司",
    "宁波亿泓国际物流有限公司",
    "深圳市中传国际物流有限公司",
    "江苏纽生堂国际货运代理有限公司",
    "江苏宏达通国际物流有限公司",
    "惠州市巨航供应链管理有限公司",
    "深圳市博洋供应链管理有限公司",
    "深圳市格威达国际货运代理有限公司",
    "常州鹏欣供应链管理有限公司",
    "深圳智中行国际物流有限公司",
    "宁波乾洋国际物流有限公司",
    "万集（宁波）供应链科技有限公司",
    "昆山贝尔通国际货运代理有限公司",
    "安徽福沣国际物流有限公司",
    "上海蕴鑫供应链管理有限公司",
    "宁波成一国际物流有限公司",
    "上海简凯国际物流有限公司",
    "安徽乐扬国际货运代理有限公司",
    "无锡万航国际货运代理有限公司",
    "浙江木牛流马供应链有限公司德清分公司",
    "上海欣乐享供应链管理有限公司",
    "宁波鼎领国际物流有限公司上海分公司",
    "神轮（苏州）国际货运代理有限公司",
    "昆山艾亘国际贸易有限公司",
    "宁波宁舟船务有限公司",
    "宁波锦海供应链有限公司",
    "江苏欣拓国际货运代理有限公司",
    "上海文鳐供应链科技有限公司",
    "宁波当曲科技有限公司",
    "浙江繁简供应链管理有限公司",
    "宁波忻辰国际供应链管理有限公司",
    "宁波集达供应链管理有限公司",
    "安吉捷泰家居有限责任公司",
    "浙江中非国际经贸港服务有限公司",
    "杭州高新供应链管理有限公司",
    "宁波海瑞得国际货运代理有限公司",
    "浙江君集国际物流有限公司",
    "广州市路瑞供应链有限公司",
    "上海融强国际物流有限公司",
    "南京唯派国际物流有限公司",
    "杭州环宇国际货运有限公司",
    "浙江天时利蜻蜓供应链科技有限公司",
    "上海青禾物流有限公司",
    "广州福施湃乐电子科技有限公司",
    "文誉供应链管理（上海）有限公司",
    "重庆创辉国际货运代理有限公司",
    "锦程国际物流集团武汉国际物流有限公司",
    "上海金阳海悦供应链管理有限公司",
    "上海诺嘉国际货物运输代理有限公司",
    "浙江至帆物流有限公司",
    "安徽卓达国际货运代理有限公司",
    "大连辉泓国际物流有限公司",
    "华基国际物流（宁波）有限公司上海分公司",
    "佛山市永正大进出口有限公司",
    "浙江嘉兴高通国际货物运输代理有限公司",
    "浙江美川国际货运代理有限公司",
    "上海衍旭物流有限公司",
    "浙江双宇电子科技有限公司",
    "安徽凯福佑国际货运代理有限公司",
    "星东国际物流（宁波）有限公司",
    "湖南西联捷运国际货运代理有限公司",
    "上海美昊杰国际货运代理有限公司",
    "上海贝莱蒂供应链管理有限公司",
    "嘉里大通物流有限公司湖北分公司",
    "宁波中合供应链有限公司",
    "宁波瑞尚供应链管理有限公司",
    "青岛明通国际物流有限公司",
    "宁波力朗供应链管理有限公司",
    "浙江跨运科技有限公司",
    "宁波柏胜国际货运代理有限公司",
    "浙江跨运供应链管理有限公司",
    "宁波甬兰国际物流有限公司",
    "宁波汇洋国际货运代理有限公司",
    "金华安睿国际货运代理有限公司",
    "吔百国际贸易（广东）有限公司",
    "宁波吉瑞国际货运代理有限公司",
    "上海中善物流有限公司",
    "锐链宁波",
    "优信嘉运国际物流（苏州）有限公司",
    "美设国际物流集团股份有限公司杭州分公司",
    "宁波久瀚国际货运代理有限公司",
    "深圳古瑞瓦特新能源有限公司",
    "上海天兆国际货物运输代理有限公司",
    "上海睿淼供应链管理有限公司",
    "领惠（上海）供应链管理有限公司",
    "上海丹吉士国际物流有限公司",
    "上海寰美国际物流有限公司",
    "上海海波物流有限公司",
    "上海复易国际货运代理有限公司",
    "融华智运国际物流（青岛）有限公司",
    "宁波熙佳国际货运代理有限公司",
    "上海大瀚国际货运代理有限公司",
    "宁波得玛供应链管理有限公司",
    "深圳市航时运通国际物流有限公司",
    "浩鸣国际物流代理有限公司",
    "嘉兴环世国际货运代理有限公司",
    "宁波寰航供应链管理有限公司",
    "诺得物流股份有限公司",
    "江苏仁义多式联运有限公司",
    "湖州安吉星辙国际货运代理有限公司",
    "湖州恒众国际货运代理有限公司",
    "浙江海舜供应链管理有限公司",
    "浙江扬程国际物流有限公司",
    "浙江绍兴美力物流有限公司",
    "博达供应链科技（宁波）有限公司",
    "南京中腾物流有限公司",
    "广州市骏亨国际货运代理有限公司",
    "浙江阿啦丁国际货运代理有限公司",
    "常州蓝湾国际货运代理有限公司",
    "义乌美星国际物流有限公司",
    "港拓（深圳）贸易有限公司",
    "上海华原国际货物运输代理有限公司",
    "诸暨市中创物流有限公司",
    "合肥远洋国际物流有限公司",
    "志存（宁波）供应链管理有限公司",
    "深圳百隆美亚国际货运代理有限公司上海分公司",
    "浙江泽惠供应链管理有限公司",
    "广州市卓志报关有限公司",
    "大连万联供应链管理有限公司",
    "宁波陆洋国际货运代理有限公司",
    "上海传鸿国际物流有限公司",
    "上海陆启通国际货物运输代理有限公司",
    "温州市鹿城区滨江晨夕货物运输服务部",
    "上海协利物流有限公司",
    "宁波田天供应链有限公司",
    "泽世国际货运代理（深圳）有限公司宁波分公司",
    "宁波美晟国际货运代理有限公司",
    "上海境港供应链科技有限公司",
    "浙江四港联动发展有限公司",
    "浙江星驰国际货运代理有限公司",
    "义乌市中超集装箱有限公司",
    "宁波伊亲国际货运代理有限公司",
    "浙江天熠国际物流有限公司",
    "浙江多多马国际货运代理有限公司",
    "湖州胜开供应链有限公司",
    "温州达丰国际物流有限公司",
    "顺圆弘通物流集团有限公司",
    "上海上楚供应链管理有限公司",
    "穰穰满仓（上海）供应链科技有限公司",
    "中集世联达亚联国际物流有限公司",
    "上海卓之烁供应链管理有限公司",
    "安徽盛凯国际货运代理有限公司",
    "丹时达供应链管理(上海)有限公司",
    "上海汇恒国际物流有限公司",
    "宁波鑫淼国际物流有限公司",
    "安徽即联即用科技有限公司",
    "宁波希澜供应链管理有限公司",
    "宁波市焕航物流有限公司",
    "福建港运国际物流有限公司",
    "青岛万家船务有限公司",
    "浙江运到数字物流发展有限公司",
    "上海樱美供应链管理有限公司",
    "嘉德运通（福州）国际货运代理有限公司厦门分公司",
    "浙江纵深供应链管理有限公司",
    "中山市海迅捷运国际物流有限公司",
    "天津瑞洋供应链管理服务有限公司",
    "宁波观龙供应链管理有限公司",
    "港联捷（上海）物流科技有限公司",
    "安徽兆洋国际物流有限公司",
    "上海海木国际物流有限公司",
    "安徽烁道国际货运代理有限公司",
    "宁波华沃供应链科技有限公司",
    "浙江新嘉国际物流有限公司",
    "杭州海杭境通供应链管理有限公司",
    "浙江海金国际货运代理有限公司",
    "上海舒雯国际物流有限公司",
    "上海粤立申国际物流有限公司",
    "宁波安洲国际物流有限公司",
    "义乌友航国际货运代理有限公司",
    "义乌市大程国际货运代理有限公司",
    "宁波万时国际物流有限公司",
    "深圳市顺云峰国际货运代理有限公司",
    "宁波中沃日新国际物流有限公司",
    "海宁吉佑佳国际货运代理有限公司",
    "浙江喜航供应链有限公司",
    "浙江振昊供应链管理有限公司",
    "上海飞远物流有限公司",
    "杭州卓远国际物流有限公司",
    "宁波锦然国际物流有限公司",
    "宁波速帆供应链管理有限公司",
    "上海秀登供应链管理有限公司",
    "宁波捷义国际物流有限公司",
    "盟湾国际物流（上海）有限公司",
    "浙江华世货运代理有限公司",
    "宁波淳宏供应链管理有限公司",
    "安徽安商国际供应链管理有限公司",
    "宁波涵瑞供应链有限公司",
    "广东航桥国际物流股份有限公司上海分公司",
    "浙江智杰国际货运代理有限公司",
    "温州灏璟国际货运代理有限公司",
    "湖北快航国际物流有限公司",
    "宁波乔山国际货运代理有限公司",
    "深圳市运通报关有限公司",
    "上海优一克国际物流有限公司",
    "信达海新国际物流（大连）有限公司",
    "浙江纺都国际货运代理有限公司",
    "衢州市捷海国际货运代理有限公司",
    "苏州德豹供应链有限公司",
    "福励物流科技（上海）有限公司",
    "上海嘉其国际物流有限公司",
    "南京上楚供应链有限责任公司",
    "宁波天航国际物流有限公司",
    "上海茂曦轩国际物流有限公司",
    "宁波锦昌国际物流有限公司",
    "宁波捷时进出口有限公司",
    "江苏海丰国际货运代理有限公司",
    "浙江嘉澜国际货运代理有限公司",
    "江苏海航供应链管理服务有限公司",
    "上海大展国际物流有限公司",
    "上海升笙国际物流有限公司",
    "海邦（江苏）国际物流有限公司",
    "浙江海创国际货运代理有限公司",
    "江苏航吉国际货运代理有限公司",
    "宁波盛威国际物流有限公司",
    "宁波联洋船务有限公司",
    "宁波民亨国际物流有限公司",
    "湖州易航国际货运代理有限公司",
    "宁波大千国际物流有限公司",
    "宁波路遥国际货运代理有限公司",
    "浙江美升国际货运代理有限公司",
    "宁波伟大联盟国际货运代理有限公司",
    "杭州达升坤道国际货运代理有限公司",
    "天津安海物流发展有限公司上海分公司",
    "浙江锦鸿物流有限公司",
    "浙江华凯供应链管理有限公司",
    "浙江佰度物流有限公司",
    "浙江经茂国际货运代理有限公司",
    "浙江韵昇航运有限公司",
    "浙江福瑞鑫供应链服务有限公司",
    "浙江凤起物流有限公司",
    "昆山邦达国际货运代理有限公司",
    "宁波鸿远国际物流有限公司",
    "浙江翼港物流有限公司",
    "宁波港泰国际物流有限公司",
    "温州承运国际货运有限公司",
    "浙江湖州华凯国际货运代理有限公司",
    "宁波亦豪供应链管理有限公司",
    "义乌市源辉物流有限公司",
    "浙江耐斯特国际货运代理有限公司",
    "宁波竭诚国际物流有限公司",
    "宁波顺圆物流有限公司",
    "宁波中景国际货运代理有限公司",
    "宁波博皓国际物流有限公司",
    "宁波杰兰德国际物流有限公司",
    "宁波优诺舟国际货运代理有限公司",
    "浙江融易通企业服务有限公司",
    "宁波希朗国际货运代理有限公司",
    "宁波源联国际货运代理有限公司",
    "杭州华意国际货运代理有限公司",
    "宁波星赛蒂国际货物运输代理有限公司",
    "浙江海盟国际货运代理有限公司",
    "嘉兴祺达国际货运代理有限公司",
    "杭州泛远国际物流股份有限公司",
    "宁波志诺国际物流有限公司",
    "浙江鑫骋物流有限公司",
    "杭州金海岸国际货运代理有限公司",
    "港捷（中国）国际货运有限公司宁波分公司",
    "江苏天禄国际货运代理有限公司",
    "浙江开瑞国际物流有限公司",
    "杭州和邦国际货运代理有限公司",
    "浙江锦睿国际货运代理有限公司",
    "上海万年青国际物流有限公司宁波分公司",
    "上海奥集国际货运代理有限公司",
    "西格玛国际物流（上海）有限公司",
    "浙江三石力供应链管理有限公司",
    "浙江天闳国际物流有限公司",
    "上海义统国际物流有限公司",
    "上海康芸物流发展有限公司",
    "苏州鑫梦成物流有限公司",
    "锦程国际物流集团上海国际物流有限公司",
    "嘉兴伟诚国际货运代理有限公司",
    "上海家林国际货物运输代理有限公司",
    "大连海升国际物流有限公司",
    "浙江禹洋国际货运代理有限公司",
    "上海美凰国际物流有限公司",
    "大连汇诚睿国际物流有限公司",
    "海盟控股集团有限公司",
    "宁波骝马供应链管理有限公司",
    "浙江长川物流有限公司",
    "宁波和瀚国际货运代理有限公司",
    "江苏嘉格国际货物运输代理有限公司",
    "上海诺亚国际物流有限公司",
    "宁波世纪泊洋国际物流有限公司",
    "湖州振和供应链管理有限公司",
    "江苏新天国际物流有限公司",
    "嘉兴麒成供应链有限公司",
    "上海聚翰船务有限公司",
    "上海富天国际物流有限公司",
    "宁波乐洲国际物流有限公司",
    "浙江恒圣供应链管理有限公司",
    "宁波市联合大洋物流有限公司",
    "上海全运货运代理有限公司",
    "宁波骏联国际货运代理有限公司",
    "江苏鑫博亚国际物流有限公司",
    "飞力达供应链（上海）有限公司",
    "天徽国际供应链管理有限公司",
    "南通金帆国际货运代理有限公司",
    "常熟市华盛国际货运代理有限公司",
    "安徽渡达国际货运代理有限公司",
    "杭州至星国际货运代理有限公司",
    "安徽海汐国际货运有限公司",
    "上海枚一国际货物运输代理有限公司",
    "深圳市集万国际货物运输有限公司",
    "青岛舜安恒泰国际物流有限公司杭州分公司",
    "上海羽嘉国际物流有限公司",
    "格瑞德国际货运代理（苏州）有限公司",
    "沃泰国际物流（宁波）有限公司",
    "一席货运代理（金华）有限公司",
    "上海久瀚国际货运代理有限公司",
    "上海嘉赫国际物流有限公司",
    "浙江胜马国际货运代理有限公司",
    "中宏飞集国际货运代理有限公司",
    "浙江锦诚国际货运代理有限公司",
    "上海新新运国际物流有限公司杭州分公司",
    "嘉兴中航国际货运代理股份有限公司",
    "湖州天航国际物流有限公司",
    "上海恒扬国际物流有限公司",
    "嘉兴恒元国际货运代理有限公司",
    "上海冠农国际物流有限公司",
    "杭州港翔致昇国际货运代理有限公司",
    "湖北航嘉国际物流有限公司",
    "嘉兴锦诚国际物流有限公司",
    "临海市正大国际货运代理有限公司",
    "宁波美速嘉运国际物流有限公司",
    "宁波终赢供应链管理有限公司",
    "宁波风速供应链有限公司",
    "宁波环翎国际货运代理有限公司",
    "宁波鹏信国际货运代理有限公司",
    "宁波腾海供应链管理有限公司",
    "宁波天拓泰世通国际货运代理有限公司",
    "宁波腾辉国际物流有限公司",
    "宁波凯腾国际物流有限公司",
    "宁波无双供应链管理有限公司",
    "宁波亿天国际货运代理有限公司",
    "宁波外运国际集装箱货运有限公司",
    "宁波星皓国际物流有限公司",
    "宁波路遥供应链科技（集团）有限公司",
    "优尼科供应链（深圳）有限公司宁波分公司",
    "仕巴伊努(杭州)供应链有限公司",
    "昇冠(宁波)国际货运代理有限公司",
    "上海金色物流有限公司",
    "上海鸿茂国际物流有限公司宁波分公司",
    "绍兴海隆国际货运代理有限公司",
    "深圳市顺和盈货运代理有限公司",
    "深圳市辰帆国际货运代理有限公司",
    "深圳翰远国际物流有限公司",
    "天津东时供应链管理有限公司",
    "台州源广达迅国际货运代理有限公司",
    "咏耀船务（青岛）有限公司",
    "中博润达物流(浙江)有限公司",
    "宁波中创远合物流有限公司",
    "青岛慧通世达国际物流有限公司",
    "宁波君桥国际物流有限公司",
    "宁波金仑国际物流有限公司",
    "宁波怡升昌供应链管理有限公司",
    "宁波终赢供应链管理有限公司",
    "上海恒汇达国际物流有限公司",
    "绍兴柯桥红胜火纺织品有限公司",
    "浙江海亮股份有限公司",
    "宁波开万国际物流有限公司",
    "青岛山建商贸有限公司",
    "卓辰实业（上海）有限公司",
    "宁波天美利进出口有限公司",
    "宁波东曜电器有限公司",
    "舟山晟利食品有限公司",
    "西格玛国际物流（上海）有限公司",
    "宁波来回转自动化工业有限公司",
    "浙江海亮股份有限公司",
    "中国邮政速递物流股份有限公司绍兴市分公司",
    "杭州凤起国际货运代理有限公司",
    "浙江你好国际物流有限公司",
    "中国邮政速递物流股份有限公司衢州市分公司",
    "浙江外代国际物流有限公司",
    "浙江启胜国际货运代理有限公司",
    "浙江聚航国际货运代理有限公司",
    "浙江港盈科技有限公司",
    "浙江金斯顿供应链集团有限公司",
    "浙江省桐乡经济开发区管理委员会",
    "浙江易鑫国际货运代理有限公司",
    "智美通国际物流(深圳)有限公司",
    "众翔供应链管理(浙江)有限公司",
    "易豹网络科技有限公司",
    "浙江海蛛网络科技有限公司",
    "浙江运到数字物流发展有限公司"
]

# if __name__ == '__main__':
#     # print("输出几个统计维度的时间")
#     # # 输出本月第一天
#     print(this_month_start)
#     # 输出上个月第一天
#     print(this_last_month_start)
#     # # # 输出本月最后一天
#     # print(this_month_end)
#     # # # 输出今年开年第一天
#     print(this_year_start)
#     # print("获取所有目标客户公司系统内唯一 Id 标记")
#     idList = idCapture(companies)
#     # print(idList)
#     print("大笨鸟拖车/大笨鸟报关/大笨鸟操作票数/小笨鸟报关/小笨鸟拖车")
# idList = idCapture(company_list)

#     for i in range(len(idList)):
#         print(companies[i])
#         timeFlagA = "2025-04-01 00:00:00"
#         # timeFlagA = this_year_start
#         timeFlagB = "2025-05-01 00:00:00"
#
#         print("大笨鸟拖车")
#         dabenniaotuoche(idList[i], timeFlagA, timeFlagB)
#
#         print("大笨鸟报关")
#         dabenniaobaoguan(idList[i], timeFlagA, timeFlagB)
#
#         print("大笨鸟操作票数")
#         dabenniaocaozuopiaoshu(idList[i], timeFlagA, timeFlagB)
#
#         print("小笨鸟报关")
#         xiaobenniaobaoguan(idList[i], timeFlagA, timeFlagB)
#
#         print("小笨鸟拖车")
#         xiaobenniaotuoche(idList[i], timeFlagA, timeFlagB)
#
#         print("*******************************************************************")
#
#     print("统计完成")
#     # print("大笨鸟操作票数列表")
#
#     print(len(listSeqA))
#     print(listSeqA)
#
#     print(len(listSeqB))
#     print(listSeqB)
#
#     print(len(listSeqC))
#     print(listSeqC)
#
#     print(len(listSeqD))
#     print(listSeqD)
#
#     print(len(listSeqE))
#     print(listSeqE)
#
#     # 单元格合并测试代码
#     from openpyxl import Workbook
#
#     # 创建一个新的工作簿
#     wb = Workbook()
#
#     # 获取当前活动的工作表
#     ws = wb.active
#
#     # 重命名工作表
#     ws.title = "大笨鸟数据统计"
#     ws.merge_cells(start_row=5, start_column=3, end_row=6, end_column=3)
#     ws.merge_cells(start_row=7, start_column=3, end_row=8, end_column=3)
#     ws.merge_cells(start_row=9, start_column=3, end_row=10, end_column=3)
#     ws.merge_cells(start_row=11, start_column=3, end_row=12, end_column=3)
#
#     ws.merge_cells(start_row=13, start_column=3, end_row=17, end_column=3)
#
#     ws.merge_cells(start_row=18, start_column=3, end_row=19, end_column=3)
#     ws.merge_cells(start_row=20, start_column=3, end_row=21, end_column=3)
#     ws.merge_cells(start_row=22, start_column=3, end_row=23, end_column=3)
#     ws.merge_cells(start_row=24, start_column=3, end_row=25, end_column=3)
#     ws.merge_cells(start_row=26, start_column=3, end_row=27, end_column=3)
#     ws.merge_cells(start_row=28, start_column=3, end_row=29, end_column=3)
#     ws.merge_cells(start_row=30, start_column=3, end_row=31, end_column=3)
#
#     ws.merge_cells(start_row=32, start_column=3, end_row=39, end_column=3)
#
#     for i in range(40, 99, 2):
#         ws.merge_cells(start_row=i, start_column=3, end_row=i + 1, end_column=3)
#
#     for p in range(1, 5):
#         ws['C' + str(5 + (p - 1) * 2)] = listSeqE[p - 1]
#
#     ws['C13'] = listSeqE[4] + listSeqE[5] + listSeqE[6] + listSeqE[7] + listSeqE[8]
#
#     for p in range(8, 15):
#         ws['C' + str(4 + (p - 1) * 2)] = listSeqE[p + 1]
#
#     ws['C32'] = listSeqE[16] + listSeqE[17] + listSeqE[18] + listSeqE[19] + listSeqE[20] + listSeqE[21] + listSeqE[22] + \
#                 listSeqE[23]
#
#     for p in range(18, 45):
#         ws['C' + str(6 + (p - 1) * 2)] = listSeqE[p + 5]
#
#         # 写入数据到单元格
#     # ws['C5'] = 'Top Left'
#     # ws['C7'] = 'Top Right'
#
#     # ws['A2'] = 'Bottom Left'
#     # ws['C2'] = 'Bottom Right'
#     # 保存工作簿到文件
#     wb.save("大笨鸟数据统计.xlsx")


def xuanniao(theids, timeFlagA, timeFlagB):
    # user = 'dubtest'
    # passwd = 'Di3Vo#eB'
    # port = 3309
    # db = 'lit_saas_booking'
    # host = "192.168.17.12"


    user = 'litsaas'
    passwd = 'Litsaas_0607'
    port = 3306
    db = 'litsaas'
    host = "rm-uf6160k88d3snm052.mysql.rds.aliyuncs.com"



    # 连接数据库
    conn = pymysql.connect(host=host  # 连接名称，默认127.0.0.1
                           , user=user  # 用户名
                           , passwd=passwd  # 密码
                           , port=port  # 端口，默认为3306
                           , db=db  # 数据库名称
                           , charset='utf8'  # 字符编码
                           )
    cur = conn.cursor()  # 生成游标对象

    # SELECT
    # COUNT(*)
    # AS
    # total
    # FROM
    # lit_job_head
    # WHERE
    # biz_type
    # IN('SE', 'SI')
    # AND
    # cust_name = '火焰公司'
    # AND
    # deleted_flag = '0';

    # try:
        # sql = "insert into ocean_flight_etd_eta3  values(%s,%s,%s,%s,%s)"
        # # sql = "insert into" + table1 + "values(%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)"
        # param = (a,b,c,d,e)
        # mycursor = mydb.cursor()
    # sql = """
    # SELECT ORG_ID,count(1) FROM lit_booking_declare_order WHERE DELETED_FLAG = '0' and declare_way = '0' and org_id in (%(theid)s) and create_time >= '2024-03-01 00:00:00' and create_time < '2023-09-01 00:00:00' GROUP BY ORG_ID;
    # """
    #     sql = """SELECT  job_no FROM `lit_job_head` WHERE  deleted_flag = 0 and job_id = 1407978071168032777;
    #     SELECT ORG_ID,count(1) FROM lit_booking_declare_order WHERE DELETED_FLAG = '0' and declare_way = '0' and org_id in (%(theid)s) and create_time >= '2024-03-01 00:00:00' and create_time < '2023-09-01 00:00:00' GROUP BY ORG_ID;
    #     """
    sql = """
    
    
    
    SELECT COUNT(*) AS total
FROM lit_job_head
WHERE biz_type IN ('SE', 'SI')
  AND cust_name = %(theid)s
  AND deleted_flag = '0';
    """
    # cur.execute(sql)
    cur.execute(sql, {"theid": theids})
    # cur.execute("")
    myresult = cur.fetchall()

    if myresult:
        for x in myresult:
            print(theids,x)  # 打印每条记录
            excel_data.append((theids, x[0]))
    else:
        print(theids, (0,))
        excel_data.append((theids, 0))

        # cur.execute(sql, param)
        # conn.commit() # 提交到数据库执行
    # except Exception as e:
    #     print(e)
    #     conn.rollback()  # 如果发生错误则回滚
    #     conn.close()  # 关闭数据库连接


# if __name__ == '__main__':
#
#     # print("输出几个统计维度的时间")
#     # # 输出本月第一天
#     print(this_month_start)
#     # 输出上个月第一天
#     print(this_last_month_start)
#     # # # 输出本月最后一天
#     # print(this_month_end)
#     # # # 输出今年开年第一天
#     print(this_year_start)
#     # print("获取所有目标客户公司系统内唯一 Id 标记")
#     # idList = idCapture(companies)
#     # print(idList)
#     print("玄鸟订舱客户")
#     for i in range(0,1):
#         # print(companies[i])
#         timeFlagA = "2020-01-01 00:00:00"
#         # timeFlagA = this_year_start
#         timeFlagB = "2025-09-01 00:00:00"
#
#         print("玄鸟")
#         xuanniao ("中集世联达国际供应链有限公司", timeFlagA, timeFlagB)
#
#         print("*******************************************************************")
if __name__ == '__main__':
    # 打开文件并重定向print输出
    # with open('输出结果.txt', 'w', encoding='utf-8') as f:
    #     # 保存原始的print函数
    #     original_print = print
    #
    #
    #     # 定义新的print函数，同时输出到控制台和文件
    #     idList = (company_list)
    #     def print_to_file(*args, **kwargs):
    #         original_print(*args, **kwargs)  # 正常打印到控制台
    #         # 将内容写入文件
    #         f.write(' '.join(str(arg) for arg in args) + '\n')
    #         f.flush()  # 立即写入文件
    #
    #
    #     # 替换print函数
    #     import builtins
    #
    #     builtins.print = print_to_file
    #
    #     # 您现有的代码...
    #     # print("输出几个统计维度的时间")
    #     # print(this_month_start)
    #     # print(this_last_month_start)
    #     # print(this_year_start)
    #     # print("玄鸟订舱客户")
    #
    #     for i in range(0, len(idList)):
    #         timeFlagA = "2020-01-01 00:00:00"
    #         timeFlagB = "2025-09-01 00:00:00"
    #         print("玄鸟")
    #         xuanniao(idList[i], timeFlagA, timeFlagB)
    #         print("*******************************************************************")
    #
    #     # 恢复原始print函数
    #     builtins.print = original_print

    excel_data = []  # 重新初始化列表
    idList = (company_list)

    for i in range(0, len(idList)):
        timeFlagA = "2020-01-01 00:00:00"
        timeFlagB = "2025-09-01 00:00:00"
        print("玄鸟")
        xuanniao(idList[i], timeFlagA, timeFlagB)
        print("*******************************************************************")
    # for i in range(0, len(idList)):
    # timeFlagA = "2020-01-01 00:00:00"
    # timeFlagB = "2025-09-01 00:00:00"
    # print("玄鸟")
    # xuanniao('saila', timeFlagA, timeFlagB)
    # print("*******************************************************************")

    # 将数据写入Excel文件
    if excel_data:
        df = pd.DataFrame(excel_data, columns=['公司名称', '数量'])
        df.to_excel('玄鸟统计结果.xlsx', index=False)
        print(f"数据已保存到Excel文件，共{len(excel_data)}条记录")